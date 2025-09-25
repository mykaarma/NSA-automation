import os
import requests
import time
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from dotenv import load_dotenv
import openpyxl
from openpyxl import Workbook
from dealer_info import DEALERS, get_dealer_by_id
import json
from communication_service import create_communication_service

try:
    from tqdm import tqdm
    USE_TQDM = True
except ImportError:
    USE_TQDM = False

# Load environment variables
load_dotenv()

BASE_URL = os.getenv('MYKAARMA_BASE_URL')
USERNAME = os.getenv('MYKAARMA_USERNAME')
PASSWORD = os.getenv('MYKAARMA_PASSWORD')
AUTH = (USERNAME, PASSWORD)

# Communication settings
SEND_TEXT_NOTIFICATIONS = True
SEND_EMAIL_NOTIFICATIONS = True

# Cache settings
CACHE_FILE = 'appointment_cache.json'

def load_appointment_cache():
    """
    Load the appointment cache from file.
    
    Returns:
        Dictionary containing cached appointment data
    """
    try:
        if os.path.exists(CACHE_FILE):
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            return {"cached_orders": []}
    except Exception as e:
        print(f"Warning: Could not load appointment cache: {e}")
        return {"cached_orders": []}

def save_appointment_cache(cache_data):
    """
    Save the appointment cache to file.
    
    Args:
        cache_data: Dictionary containing cache data to save
    """
    try:
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Warning: Could not save appointment cache: {e}")

def is_order_cached(cache_data, ro_number):
    """
    Check if an RO number is already in the cache.
    
    Args:
        cache_data: Cache dictionary
        ro_number: RO number to check
        
    Returns:
        Cached order entry if found, None otherwise
    """
    for cached_order in cache_data.get("cached_orders", []):
        if cached_order.get("ro_number") == ro_number:
            return cached_order
    return None

def add_to_cache(cache_data, ro_number, customer_first_name, customer_last_name, dealer_id, appointment_uuid):
    """
    Add a successfully created appointment to the cache.
    
    Args:
        cache_data: Cache dictionary
        ro_number: RO number
        customer_first_name: Customer's first name
        customer_last_name: Customer's last name
        dealer_id: Dealer ID
        appointment_uuid: Created appointment UUID
    """
    cache_entry = {
        "ro_number": ro_number,
        "customer_first_name": customer_first_name,
        "customer_last_name": customer_last_name,
        "dealer_id": dealer_id,
        "created_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "appointment_uuid": appointment_uuid
    }
    
    # Remove any existing entry for this RO number (shouldn't happen, but just in case)
    cache_data["cached_orders"] = [entry for entry in cache_data.get("cached_orders", []) 
                                   if entry.get("ro_number") != ro_number]
    
    # Add the new entry
    cache_data["cached_orders"].append(cache_entry)

def check_for_duplicates(rows, cache_data):
    """
    Check for duplicate RO numbers in the cache.
    
    Args:
        rows: List of rows to process
        cache_data: Cache dictionary
        
    Returns:
        List of duplicate entries found
    """
    duplicates = []
    
    for row in rows:
        ro_number = row.get('RO Number')
        if ro_number:
            cached_entry = is_order_cached(cache_data, ro_number)
            if cached_entry:
                duplicates.append({
                    "ro_number": ro_number,
                    "customer_name": f"{row.get('Customer First Name', '')} {row.get('Customer Last Name', '')}".strip(),
                    "dealer_id": row.get('Dealer ID'),
                    "cached_date": cached_entry.get('created_date'),
                    "cached_appointment_uuid": cached_entry.get('appointment_uuid')
                })
    
    return duplicates

def prompt_user_for_duplicates(duplicates):
    """
    Prompt user about duplicate RO numbers and ask if they want to create new appointments.
    
    Args:
        duplicates: List of duplicate entries
        
    Returns:
        True if user wants to continue, False otherwise
    """
    print("\n" + "="*60)
    print("DUPLICATE RO NUMBERS DETECTED")
    print("="*60)
    print(f"Found {len(duplicates)} RO(s) that already have appointments created:")
    print()
    
    for i, dup in enumerate(duplicates, 1):
        print(f"{i}. RO Number: {dup['ro_number']}")
        print(f"   Customer: {dup['customer_name']}")
        print(f"   Dealer ID: {dup['dealer_id']}")
        print(f"   Previous appointment created: {dup['cached_date']}")
        print(f"   Previous appointment UUID: {dup['cached_appointment_uuid']}")
        print()
    
    while True:
        response = input("Do you want to create new appointments for these ROs anyway? (y/n): ").strip().lower()
        if response in ['y', 'yes']:
            print("Continuing to create new appointments...")
            return True
        elif response in ['n', 'no']:
            print("Skipping appointment creation for duplicate ROs.")
            return False
        else:
            print("Please enter 'y' for yes or 'n' for no.")

def fetch_slot_size(dealer_uuid):
    url = f"{BASE_URL}/appointment/v2/dealer/{dealer_uuid}/hoursOfOperation"
    headers = {"accept": "application/json"}
    cookies = {"rollout.stage": "canary"}
    resp = requests.get(url, auth=AUTH, headers=headers, cookies=cookies)
    resp.raise_for_status()
    data = resp.json()
    return int(data.get('slotSizeInMins', 15))  # Default to 15 if not present

def load_opcodes_from_xlsx(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    opcodes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            opcode = str(row[0]).strip()
            description = str(row[1]).strip() if row[1] else ""
            opcodes[opcode] = description
    return opcodes

def get_first_available_slot_firstapi(row, department_uuid, filtered_opcodes, target_date):
    url = f"{BASE_URL}/appointment/v2/department/{department_uuid}/first-available-slot"
    headers = {"accept": "application/json", "Content-Type": "application/json"}
    cookies = {"rollout.stage": "canary"}
    # If target_date is in the past, use today
    from_date = max(target_date, datetime.today())
    date_str = from_date.strftime('%Y-%m-%d')
    body = {
        "dates": [date_str],
        "customerInformation": {
            "firstName": row.get('Customer First Name'),
            "lastName": row.get('Customer Last Name'),
            "uuid": row.get('Customer UUID'),
            "key": row.get('Customer Key'),
        },
        "vehicleInformation": {
            "uuid": row.get('Vehicle UUID'),
            "vin": row.get('VIN'),
        },
        "laborOpcodeList": filtered_opcodes,
        "selectedAvailabilityAttributes": {},
        "allAvailabilityAttributes": {}
    }
    resp = requests.post(url, json=body, auth=AUTH, headers=headers, cookies=cookies)
    resp.raise_for_status()
    data = resp.json()
    dt = data.get('dateTime')
    if dt:
        # dt is like '2024-08-01 09:15:00'
        appt_date, appt_time = dt.split(' ')
        # Log the first available slot received
        print(f"    [First Available Slot] RO: {row.get('RO Number', 'N/A')}, Customer: {row.get('Customer First Name', '')} {row.get('Customer Last Name', '')}, Date: {appt_date}, Time: {appt_time}")
        return appt_date, appt_time
    print(f"    [First Available Slot] No slot found for RO: {row.get('RO Number', 'N/A')}, Customer: {row.get('Customer First Name', '')} {row.get('Customer Last Name', '')}")
    return None, None

def create_appointment(row, appt_date, appt_time, slot_size, dealer_uuid, filtered_opcodes, opcode_descriptions):
    url = f"{BASE_URL}/appointment/v2/dealer/{dealer_uuid}/appointment"
    headers = {"accept": "application/json", "Content-Type": "application/json"}
    cookies = {"rollout.stage": "canary"}
    service_list = []
    for op in filtered_opcodes:
        service_item = {"title": op, "operationType": "OPCODE"}
        if opcode_descriptions.get(op):
            service_item["description"] = opcode_descriptions[op]
        service_list.append(service_item)
    # Calculate end time
    start_dt = datetime.strptime(f"{appt_date}T{appt_time}", "%Y-%m-%dT%H:%M:%S")
    end_dt = start_dt + timedelta(minutes=slot_size) - timedelta(seconds=1)
    end_time = end_dt.strftime("%H:%M:%S")
    body = {
        "customerUuid": row['Customer UUID'],
        "vehicleInformation": {
            "vehicleUuid": row['Vehicle UUID'],
            "vin": row['VIN']
        },
        "appointmentInformation": {
            "appointmentStartDateTime": f"{appt_date}T{appt_time}",
            "appointmentEndDateTime": f"{appt_date}T{end_time}",
            "serviceList": service_list,
            "comments": "",
            "internalNotes": "Next Service Appointment scheduled automatically by script.",
            "customerAppointmentPreference": {
                "emailConfirmation": True,
                "textConfirmation": True,
                "emailReminder": True,
                "textReminder": True,
                "notifyCustomer": False,
                "sendCommunicationToDA": False,
            },
            "status": None,
            "recall": False,
            "pushToDms": True
        }
    }
    resp = requests.post(url, json=body, auth=AUTH, headers=headers, cookies=cookies)
    resp.raise_for_status()
    return resp.json()

def send_appointment_notifications(communication_service, row, dealer_info, appt_date=None, appt_time=None):
    """
    Send text and email notifications for a scheduled appointment.
    
    Args:
        communication_service: CommunicationService instance
        row: Row data containing customer information
        dealer_info: Dealer information dictionary
        appt_date: Appointment date in YYYY-MM-DD format
        appt_time: Appointment time in HH:MM:SS format
        
    Returns:
        Dictionary with notification results
    """
    if not (SEND_TEXT_NOTIFICATIONS or SEND_EMAIL_NOTIFICATIONS):
        return {"status": "SKIPPED", "reason": "Notifications disabled"}
    
    try:
        # Extract customer information
        customer_firstname = row.get('Customer First Name', '')
        customer_lastname = row.get('Customer Last Name', '')
        customer_uuid = row.get('Customer UUID')
        
        # Get dealer information
        dealer_name = dealer_info.get('name', 'Our Service Center')
        department_uuid = dealer_info.get('department_uuid')
        
        if not customer_uuid:
            return {"status": "FAILED", "reason": "No customer UUID available"}
        
        if not department_uuid:
            return {"status": "FAILED", "reason": "No department UUID available"}
        
        # Send notifications - communication service will fetch default dealer associate automatically
        notification_result = communication_service.send_appointment_notifications(
            department_uuid=department_uuid,
            customer_uuid=customer_uuid,
            customer_firstname=customer_firstname,
            customer_lastname=customer_lastname,
            dealer_name=dealer_name,
            appt_date=appt_date,
            appt_time=appt_time,
            send_text=SEND_TEXT_NOTIFICATIONS,
            send_email=SEND_EMAIL_NOTIFICATIONS
        )
        
        return notification_result
        
    except Exception as e:
        return {"status": "FAILED", "error": str(e)}


def prefetch_dealer_context(rows):
    unique_dealer_ids = set(row['Dealer ID'] for row in rows)
    dealer_context = {}
    for dealer_id in unique_dealer_ids:
        dealer_info = get_dealer_by_id(dealer_id)
        if not dealer_info:
            dealer_context[dealer_id] = None
            continue
        slot_size = fetch_slot_size(dealer_info['dealer_uuid'])
        valid_opcodes = load_opcodes_from_xlsx(dealer_info['opcode_xlsx'])
        dealer_context[dealer_id] = {
            'dealer_info': dealer_info,
            'slot_size': slot_size,
            'valid_opcodes': valid_opcodes
        }
    return dealer_context

def main():
    xlsx_file = "closed_ros.xlsx"
    if not os.path.exists(xlsx_file):
        print("Error: closed_ros.xlsx not found in the current directory.")
        return
    
    # Initialize communication service
    communication_service = None
    if SEND_TEXT_NOTIFICATIONS or SEND_EMAIL_NOTIFICATIONS:
        try:
            communication_service = create_communication_service()
            print(f"Communication service initialized. Text: {SEND_TEXT_NOTIFICATIONS}, Email: {SEND_EMAIL_NOTIFICATIONS}")
        except Exception as e:
            print(f"Warning: Failed to initialize communication service: {e}")
            print("Continuing without notifications...")
    
    # Load appointment cache
    print("Loading appointment cache...")
    cache_data = load_appointment_cache()
    
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active
    rows = [dict(zip([cell.value for cell in ws[1]], [cell.value for cell in row])) for row in ws.iter_rows(min_row=2)]
    total = len(rows)
    
    # Check for duplicates before processing
    print("Checking for duplicate RO numbers...")
    duplicates = check_for_duplicates(rows, cache_data)
    
    if duplicates:
        if not prompt_user_for_duplicates(duplicates):
            print("Exiting without creating appointments.")
            return
        print("\nProceeding with appointment creation...\n")
    else:
        print("No duplicate RO numbers found. Proceeding with appointment creation...\n")
    
    dealer_context = prefetch_dealer_context(rows)
    results = []
    if USE_TQDM:
        iterator = tqdm(rows, desc="Scheduling Appointments", unit="appt")
    else:
        iterator = rows
    for idx, row in enumerate(iterator, 1):
        dealer_id = row['Dealer ID']
        context = dealer_context.get(dealer_id)
        if not context or not context['dealer_info']:
            print(f"  [{idx}/{total}] Dealer ID {dealer_id} not found in dealer_info.py. Skipping.")
            row['NSA Status'] = 'FAILED'
            row['NSA Date'] = ''
            row['NSA UUID'] = ''
            row['Text Notification Status'] = 'SKIPPED'
            row['Email Notification Status'] = 'SKIPPED'
            results.append(row)
            continue
        dealer_info = context['dealer_info']
        slot_size = context['slot_size']
        valid_opcodes = context['valid_opcodes']
        opcodes = row['Opcodes'].split(',') if row.get('Opcodes') else []
        filtered_opcodes = [op for op in opcodes if op in valid_opcodes]
        
        # Add default NSA opcode for this dealer (for reporting and identification)
        default_nsa_opcode = dealer_info.get('default_nsa_opcode')
        if default_nsa_opcode and default_nsa_opcode not in filtered_opcodes:
            filtered_opcodes.append(default_nsa_opcode)
        
        # Create opcode_descriptions dict for filtered opcodes
        opcode_descriptions = {op: valid_opcodes.get(op, '') for op in filtered_opcodes}
        close_date = row['RO Close Date']
        months = dealer_info.get('next_service_interval_in_months', 6)
        target_date = datetime.strptime(close_date, '%Y-%m-%d') + relativedelta(months=months)
        appt_uuid = None
        appt_date = None
        appt_time = None
        for attempt in range(1, 3):
            appt_date, appt_time = get_first_available_slot_firstapi(row, dealer_info['department_uuid'], filtered_opcodes, target_date)
            # Log the first available slot received (already logged in get_first_available_slot_firstapi)
            if appt_date and appt_time:
                try:
                    resp = create_appointment(row, appt_date, appt_time, slot_size, dealer_info['dealer_uuid'], filtered_opcodes, opcode_descriptions)
                    appt_uuid = resp.get('appointmentUuid', 'Success')
                    row['NSA Status'] = 'SUCCESS'
                    row['NSA Date'] = f"{appt_date} {appt_time}"
                    row['NSA UUID'] = appt_uuid
                    
                    # Add to cache after successful appointment creation
                    add_to_cache(
                        cache_data, 
                        row.get('RO Number'), 
                        row.get('Customer First Name', ''), 
                        row.get('Customer Last Name', ''), 
                        dealer_id, 
                        appt_uuid
                    )
                    
                    # Send notifications if communication service is available
                    if communication_service:
                        notification_result = send_appointment_notifications(communication_service, row, dealer_info, appt_date, appt_time)
                        
                        # Update notification status fields
                        text_result = notification_result.get('text_result', {})
                        email_result = notification_result.get('email_result', {})
                        
                        row['Text Notification Status'] = text_result.get('status', 'NOT_ATTEMPTED')
                        row['Email Notification Status'] = email_result.get('status', 'NOT_ATTEMPTED')
                        
                        if notification_result.get('overall_status') == 'SUCCESS':
                            print(f"  [{idx}/{total}] Appointment created and notifications sent: {appt_uuid}")
                        elif notification_result.get('overall_status') == 'PARTIAL_FAILED':
                            print(f"  [{idx}/{total}] Appointment created, some notifications failed: {appt_uuid}")
                        else:
                            print(f"  [{idx}/{total}] Appointment created, notifications failed: {appt_uuid}")
                    else:
                        row['Text Notification Status'] = 'DISABLED'
                        row['Email Notification Status'] = 'DISABLED'
                        print(f"  [{idx}/{total}] Appointment created for {appt_date} {appt_time}: {appt_uuid}")
                    
                    results.append(row)
                    break
                except Exception as e:
                    print(f"  [{idx}/{total}] Attempt {attempt} failed: {e}")
                    time.sleep(2)
            else:
                print(f"  [{idx}/{total}] No available slot found after {target_date.date()} (using first-available-slot endpoint). Attempt {attempt}.")
                time.sleep(2)
        if not appt_uuid:
            row['NSA Status'] = 'FAILED'
            row['NSA Date'] = ''
            row['NSA UUID'] = ''
            row['Text Notification Status'] = 'NOT_ATTEMPTED'
            row['Email Notification Status'] = 'NOT_ATTEMPTED'
            results.append(row)
        time.sleep(2)
    # Write results to xlsx
    out_filename = xlsx_file.replace('closed_ros', 'schedule_results')
    out_wb = Workbook()
    out_ws = out_wb.active
    out_fields = list(results[0].keys())
    out_ws.append(out_fields)
    for r in results:
        out_ws.append([r.get(f) for f in out_fields])
    out_wb.save(out_filename)
    
    # Save updated cache
    print("\nSaving appointment cache...")
    save_appointment_cache(cache_data)
    
    # Summary
    successful_appointments = sum(1 for r in results if r.get('NSA Status') == 'SUCCESS')
    print(f"Wrote appointment scheduling results to {out_filename}")
    print(f"\nSummary:")
    print(f"- Total rows processed: {len(results)}")
    print(f"- Successful appointments: {successful_appointments}")
    print(f"- Failed appointments: {len(results) - successful_appointments}")
    print(f"- Cache now contains: {len(cache_data.get('cached_orders', []))} total appointments")

if __name__ == "__main__":
    main() 
