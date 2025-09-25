import os
import requests
from datetime import datetime
from dotenv import load_dotenv
import openpyxl
from openpyxl import Workbook
from dealer_info import DEALERS, get_dealer_by_id
import json

try:
    from tqdm import tqdm
    USE_TQDM = True
except ImportError:
    USE_TQDM = False

# Load environment variables
load_dotenv()

BASE_URL = os.getenv('MYKAARMA_BASE_URL')
USERNAME = os.getenv('MYKAARMA_USERNAME')
PASSWORD = os.getenv('MYKAARMA_PASSWORD')
PAGE_SIZE = int(os.getenv('PAGE_SIZE', '100'))

AUTH = (USERNAME, PASSWORD)

XLSX_FIELDS = [
    'Dealer ID',
    'RO Number',
    'Order UUID',
    'Customer First Name',
    'Customer Last Name',
    'Customer Key',
    'Customer UUID',
    'VIN',
    'Vehicle UUID',
    'Opcodes',
    'RO Close Date',
    'NSA Status',
    'NSA Date',
    'NSA UUID',
]

def prompt_dealer():
    print("Select a dealer:")
    dealer_list = list(DEALERS.items())
    for idx, (k, v) in enumerate(dealer_list, 1):
        print(f"{idx}. {v['name']}")
    while True:
        try:
            sno = int(input("Enter dealer S.No.: ").strip())
            if 1 <= sno <= len(dealer_list):
                dealer_id, dealer_info = dealer_list[sno - 1]
                return dealer_id, dealer_info
        except ValueError:
            pass
        print("Invalid S.No. Try again.")

def prompt_date_range():
    """
    Prompt user to choose between single date or date range extraction.
    
    Returns:
        Tuple of (from_date, to_date, date_description)
    """
    print("\nChoose extraction method:")
    print("1. Single date")
    print("2. Date range")
    
    while True:
        try:
            choice = int(input("Enter your choice (1 or 2): ").strip())
            if choice in [1, 2]:
                break
        except ValueError:
            pass
        print("Invalid choice. Please enter 1 or 2.")
    
    if choice == 1:
        # Single date
        while True:
            date_str = input("Enter the close date (YYYY-MM-DD): ").strip()
            try:
                # Validate date format
                datetime.strptime(date_str, '%Y-%m-%d')
                return date_str, date_str, f"on {date_str}"
            except ValueError:
                print("Invalid date format. Please use YYYY-MM-DD format.")
    else:
        # Date range
        while True:
            from_date = input("Enter the start date (YYYY-MM-DD): ").strip()
            try:
                # Validate from_date format
                from_dt = datetime.strptime(from_date, '%Y-%m-%d')
                break
            except ValueError:
                print("Invalid date format. Please use YYYY-MM-DD format.")
        
        while True:
            to_date = input("Enter the end date (YYYY-MM-DD): ").strip()
            try:
                # Validate to_date format and ensure it's not before from_date
                to_dt = datetime.strptime(to_date, '%Y-%m-%d')
                if to_dt < from_dt:
                    print("End date cannot be before start date. Please try again.")
                    continue
                break
            except ValueError:
                print("Invalid date format. Please use YYYY-MM-DD format.")
        
        if from_date == to_date:
            return from_date, to_date, f"on {from_date}"
        else:
            return from_date, to_date, f"from {from_date} to {to_date}"

def load_opcodes_from_xlsx(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    opcodes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            opcodes.add(str(row[0]).strip())
    return opcodes

def fetch_closed_ros(from_date, to_date, department_uuid):
    url = f"{BASE_URL}/order/v2/department/{department_uuid}/order/specificSearch"
    headers = {"accept": "application/json"}
    body = {
        "dateFilterType": "CLOSE_DATE",
        "fromOrderDate": from_date,
        "toOrderDate": to_date,
        "orderType": "RO",
        "orderStatus": "C",
        "size": str(PAGE_SIZE)
    }
    resp = requests.post(url, json=body, auth=AUTH, headers=headers)
    resp.raise_for_status()
    return resp.json().get('orders', [])

def fetch_order_details(order_uuid, department_uuid):
    url = f"{BASE_URL}/order/v2/department/{department_uuid}/global_order/{order_uuid}"
    headers = {"accept": "application/json"}
    resp = requests.get(url, auth=AUTH, headers=headers)
    resp.raise_for_status()
    return resp.json().get('order', {})

def extract_relevant_data(order_detail, opcodes_set):
    order = order_detail.get('order', {})
    header = order.get('header', {})
    vehicle = order.get('vehicle', {})
    customer = order.get('customer', {})
    jobs = order.get('jobs', [])
    opcodes_in_ro = [job.get('laborOpCode') for job in jobs if job.get('laborOpCode')]
    if not any(op in opcodes_set for op in opcodes_in_ro):
        return None
    # Format close date to YYYY-MM-DD only
    close_date_raw = header.get('closeDate')
    ro_close_date = ''
    if close_date_raw:
        try:
            ro_close_date = datetime.strptime(close_date_raw[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
        except Exception:
            ro_close_date = close_date_raw[:10]
    return {
        'RO Number': header.get('orderNumber'),
        'Order UUID': order_detail.get('uuid'),
        'Customer First Name': customer.get('firstName'),
        'Customer Last Name': customer.get('lastName'),
        'Customer Key': customer.get('key'),
        'Customer UUID': customer.get('uuid'),
        'VIN': vehicle.get('vin'),
        'Vehicle UUID': vehicle.get('uuid'),
        'Opcodes': ','.join(opcodes_in_ro),
        'RO Close Date': ro_close_date,
        'NSA Status': '',
        'NSA Date': '',
        'NSA UUID': '',
    }

def main():
    dealer_id, dealer_info = prompt_dealer()
    xlsx_filename = f"closed_ros.xlsx"
    # Check if file exists and prompt user
    keep_existing = False
    if os.path.exists(xlsx_filename):
        print(f"{xlsx_filename} already exists.")
        ans = input(f"{xlsx_filename} already exists. Do you want to keep existing data and append to it? (y/n): ").strip().lower()
        keep_existing = (ans == 'y')
    from_date, to_date, date_description = prompt_date_range()
    opcodes_set = load_opcodes_from_xlsx(dealer_info['opcode_xlsx'])
    orders = fetch_closed_ros(from_date, to_date, dealer_info['department_uuid'])
    if orders is None:
        orders = []
    print(f"Fetched {len(orders)} closed ROs for {dealer_info['name']} {date_description}")
    rows = []
    total = len(orders)
    if USE_TQDM:
        iterator = tqdm(orders, desc="Processing ROs", unit="ro")
    else:
        iterator = orders
    for idx, o in enumerate(iterator, 1):
        order_uuid = o.get('orderUuid')
        if not order_uuid:
            continue
        try:
            detail = fetch_order_details(order_uuid, dealer_info['department_uuid'])
            row = extract_relevant_data(detail, opcodes_set)
            if row:
                row['Dealer ID'] = dealer_id
                rows.append(row)
        except Exception as e:
            print(f"Error fetching details for order {order_uuid}: {e}")
        if not USE_TQDM:
            print(f"Processed {idx} / {total} ROs", end='\r')
    # Write to XLSX
    if rows:
        if keep_existing:
            wb = openpyxl.load_workbook(xlsx_filename)
            ws = wb.active
            # Remove header row if present
            existing_headers = [cell.value for cell in ws[1]]
            if existing_headers != XLSX_FIELDS:
                ws.delete_rows(1)
                ws.insert_rows(1)
                ws.append(XLSX_FIELDS)
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(XLSX_FIELDS)
        for row in rows:
            ws.append([row.get(f) for f in XLSX_FIELDS])
        wb.save(xlsx_filename)
        print(f"\nWrote {len(rows)} filtered ROs to {xlsx_filename}")
        print(f"Date range processed: {date_description}")
    else:
        print("No ROs matched the opcode filter.")
        print(f"Date range processed: {date_description}")

if __name__ == "__main__":
    main() 